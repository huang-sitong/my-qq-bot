import asyncio
import json
from pathlib import Path

import httpx
import websockets

DOWNLOADS_DIR = Path("downloads")
DOWNLOADS_DIR.mkdir(exist_ok=True)


async def download_file(url: str, filename: str) -> Path | None:
    """下载文件到本地"""
    save_path = DOWNLOADS_DIR / filename
    if save_path.exists():
        print(f"  文件已存在，跳过: {save_path}")
        return save_path

    print(f"  下载: {url}")
    async with httpx.AsyncClient(timeout=60) as client:
        resp = await client.get(url)
        resp.raise_for_status()
        save_path.write_bytes(resp.content)
        print(f"  已保存: {save_path} ({len(resp.content)} bytes)")
        return save_path


async def process_image(file_path: Path):
    """处理图片"""
    try:
        from PIL import Image
        img = Image.open(file_path)
        print(f"  [图片] 尺寸={img.size}, 格式={img.format}, 模式={img.mode}")

        # 示例: 转为灰度图并保存
        gray = img.convert("L")
        gray_path = file_path.with_stem(file_path.stem + "_gray")
        gray.save(gray_path)
        print(f"  [图片] 灰度图已保存: {gray_path}")
    except ImportError:
        print(f"  [图片] 已下载 (需安装 Pillow 才能处理: uv add Pillow)")
    except Exception as e:
        print(f"  [图片] 处理失败: {e}")


async def process_excel(file_path: Path):
    """处理 Excel 文件"""
    try:
        import pandas as pd

        # 尝试所有 sheet
        xls = pd.ExcelFile(file_path)
        print(f"  [Excel] sheet数={len(xls.sheet_names)}, sheets={xls.sheet_names}")
        for sheet in xls.sheet_names:
            df = pd.read_excel(file_path, sheet_name=sheet)
            print(f"  [Excel] sheet[{sheet}]: {df.shape[0]}行 x {df.shape[1]}列")
            print(f"    列名: {list(df.columns)}")
            print(f"    前2行:\n{df.head(2).to_string(index=False)}")
    except ImportError:
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            print(f"  [Excel] sheet数={len(wb.sheetnames)}, sheets={wb.sheetnames}")
            for name in wb.sheetnames:
                ws = wb[name]
                print(f"  [Excel] sheet[{name}]: {ws.max_row}行 x {ws.max_column}列")
            wb.close()
        except ImportError:
            print(f"  [Excel] 已下载 (需安装 pandas/openpyxl 才能解析: uv add pandas openpyxl)")
    except Exception as e:
        print(f"  [Excel] 处理失败: {e}")


async def handle_message(body: dict):
    """处理收到的消息事件"""
    msg = body.get("message", {})
    content = msg.get("content", [])
    channel = body.get("channel", {})
    user = body.get("user", {})

    for seg in content:
        seg_type = seg.get("type")
        src = seg.get("src") or seg.get("url", "")

        if seg_type == "img":
            print(f"\n收到图片: channel={channel.get('id')}, user={user.get('id')}")
            file_path = await download_file(src, src.split("/")[-1].split("?")[0] or "image.png")
            if file_path:
                await process_image(file_path)

        elif seg_type == "file":
            title = seg.get("title", src.split("/")[-1].split("?")[0] or "file")
            print(f"\n收到文件: {title}, channel={channel.get('id')}, user={user.get('id')}")
            file_path = await download_file(src, title)
            if file_path and title.endswith((".xls", ".xlsx")):
                await process_excel(file_path)
            elif file_path:
                print(f"  [文件] 已下载，暂未注册该类型处理器: {title}")


async def main():
    uri = "ws://localhost:5600/v1/events"

    async with websockets.connect(uri) as ws:
        print(f"已连接: {uri}")

        # 鉴权
        await ws.send(json.dumps({"op": 3}))
        print("鉴权已发送 (op: 3)")

        # 持续接收事件
        async for raw in ws:
            data = json.loads(raw)
            op = data.get("op")

            if op == 0:  # 事件推送
                body = data.get("body", {})
                event_type = body.get("type")

                if event_type in ("message-created", "message"):
                    await handle_message(body)
                else:
                    print(f"\n[其他事件] type={event_type}")

            elif op == 1:  # Ping
                await ws.send(json.dumps({"op": 2}))  # 回复 Pong
            else:
                print(f"\n[其他OP] op={op} data={json.dumps(data, ensure_ascii=False)}")


asyncio.run(main())
